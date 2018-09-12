import os
import openpyxl
from openpyxl import load_workbook

#http://radiusofcircle.blogspot.com/2016/03/the-xlrd-python-module-for-reading-data.html
#https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

book = load_workbook(filename='C:/Users/Drew/PycharmProjects/excelTest/stuff.xlsx', read_only=True)
sheet = book['Samples']

i = 1

for row in sheet.iter_rows('B:B'):
    for cell in row:
        print(cell.value)